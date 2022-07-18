#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on 

@author: lars
"""

###########################_Einbindung_import_#################################
###############################################################################
import time
import datetime

import socket
import platform
import os
import subprocess
import threading
#import xlsxwriter
#import xlwt
import csv
import openpyxl

##Einbindung der Module
import sRam
import main

###########################_Funktionen_########################################
###############################################################################


################################_THREAD_#######################################

def func_th_6_thread(list,string):
    print(string)
    beginn = True

    lesen_RamSec6 = sRam.RamSec()
    schreiben_RamSec6 = sRam.RamSec()
    data_Zwischerspeicher6 = sRam.Zwischenspeicher()
    path6 = sRam.Path()

   
    time.sleep(2)
    ##Endlosschleife (verlassen nur durch Abrechen)
    while beginn == True:
        ######################################################################################
        ## Daten aus der csv-Datei in die Zwischenspeicher (Klasse -- List)
        #  
        main.semaphor_sRam_Sema.acquire()    ##Dekrementiert -1
        if lesen_RamSec6.start[0] == 1 and lesen_RamSec6.worksheet[0] == 6:
            main.semaphor_sRam_Sema.release()    ##Inkrementiert +1
            #Daten aus der CSV auslesen
            try:
                with open('AnalyseDaten\RLZ_6.csv') as csvdatei:
                    print("CSV-Datei auslesen!" + '-' * 60)
                    csv_reader_object = csv.reader(csvdatei, delimiter=';')
                    row_csv = 0
                    column_csv = 0
                    data_Zwischerspeicher6.funcClear()
                    for row in csv_reader_object:
                        #print(len(row))
                        #print(row)   ##Inhalt CSV --##Zu Ansicht einkommentieren
                        if len(row) > 1 and row[1] != "":
                            data_Zwischerspeicher6.funcSpeicher0(row[1])  #Zeile Inhalt 2.Spalte
                            row_csv += 1
                    #beginn == False
                    #break
                    print(" ")
                    print(len(data_Zwischerspeicher6.data_csv_i0))
                    print(row_csv)
                    print(" ")

            except Exception as speichern:
                print("Datei aus CSV-Datei auslesen fehlgeschlagen!" + '-' * 60)
                print(speichern)
                beginn = False
                break
        main.semaphor_sRam_Sema.release()    ##Inkrementiert +1 

        ######################################################################################
        ## Daten aus dem Zwischenspeicher (Klasse -- List) in die RLZ-Auswertung (Excel) schreiben
        #     
        main.semaphor_sRam_Sema.acquire()    ##Dekrementiert -1
        if lesen_RamSec6.start[0] == 1 and lesen_RamSec6.worksheet[0] == 6:
            main.semaphor_sRam_Sema.release()    ##Inkrementiert +1
            print("Thread_6 Analysedaten_xlsx verarbeiten, in Tabelle einfuegen und speichern!" + '-' * 60)
            try:
                print("Excel-Datei oeffnen" + '-' * 60)
                file = path6.ablagePath
                #file = 'Result_Gesamt_Analyse_RLZ.xlsx'
                fileXLSX = openpyxl.load_workbook(file)
                sheet = fileXLSX["Analyse RLZ 23° #3_1 EEH-"]
                #print(sheet['C4'].value)

                #Rechner/User-Name
                name = os.getlogin()
                sheet.cell(row=77, column=2).value = name
                rechnername = socket.gethostname()
                sheet.cell(row=77, column=3).value = rechnername
                #Datum der Auswertung
                date = datetime.datetime.now()
                sheet.cell(row=78, column=2).value = date

                print("Excel-Datei befuellen" + '-' * 60)
                zeile_xlmx = 7
                spalte_xlmx = 3
                zeile_csv = 0
                for i in range(len(data_Zwischerspeicher6.data_csv_i0)):
                    if data_Zwischerspeicher6.data_csv_i0[i] == "A" :
                        pass
                        #print("A " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "B":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("B " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "C":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("C " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "D":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("D " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "E":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("E " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "F":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("F " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "G":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("G " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "H":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("H " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "I":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("I " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "J":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("J " + '-' * 60)
                    if data_Zwischerspeicher6.data_csv_i0[i] == "K":
                        spalte_xlmx += 1
                        zeile_xlmx = 7
                        zeile_csv = 0
                        #print("K " + '-' * 60)

                    
                    if zeile_csv > 0:
                        #print(type(data_Zwischerspeicher6.data_csv_i0[i]))
                        data_Zwischerspeicher6.data_csv_i0[i]=float(data_Zwischerspeicher6.data_csv_i0[i])#.replace(".", ",")
                        #print(type(data_Zwischerspeicher6.data_csv_i0[i]))
                        sheet.cell(row=zeile_xlmx, column=spalte_xlmx).value = data_Zwischerspeicher6.data_csv_i0[i]
                        zeile_xlmx += 1
                        if zeile_xlmx == 14:
                            zeile_xlmx = 16
                        if zeile_xlmx == 29:
                            zeile_xlmx = 30
                        if zeile_xlmx == 31:
                            zeile_xlmx = 32
                        if zeile_xlmx == 43:
                            zeile_xlmx = 44
                        if zeile_xlmx == 54:
                            zeile_xlmx = 57
                    zeile_csv += 1



            except Exception as befuellen:
                print("Befuellen der Daten in Excel fehlgeschlagen!" + '-' * 60)
                print(befuellen)
                beginn = False
                break

            ######################################################################################
            ## Auswertung speichern        
            try:
                ##Excel-Datei speichern
                fileXLSX.save(path6.ablagePath)
                #fileXLSX.save('Result_Gesamt_Analyse_RLZ.xlsx')
                print("Excel-Datei gespeichert" + '-' * 60)
                
                time.sleep(1)

                main.semaphor_sRam_Sema.acquire()    ##Dekrementiert -1
                schreiben_RamSec6.funcClear()
                schreiben_RamSec6.funcClearWorksheet()
                schreiben_RamSec6.funcSec(0,1,1)
                schreiben_RamSec6.funcWS(0)
                main.semaphor_sRam_Sema.release()    ##Inkrementiert +1

                print("Eintrag fertig" + '-' * 60)
                beginn == False
                break

            except Exception as speichern:
                print("Speichern der Messung in Excel fehlgeschlagen!" + '-' * 60)
                print(speichern)
                beginn = False
                break
        main.semaphor_sRam_Sema.release()    ##Inkrementiert +1        

        if lesen_RamSec6.beenden[0] == 1:
            time.sleep(1)
            beginn = False
            break
            #sys.exit()

        # else:
        #     ##Start kann nicht durchgeführt werden
        #     print("Thread_6 Daten loggen nicht hergestellt")
        #     if lesen_RamSec6.beenden[0] == 1:
        #         beginn = False
        #         break
        #         #sys.exit()

    print("Thread_6 Analysedaten_xlsx wird beendet!" + '-' * 60)